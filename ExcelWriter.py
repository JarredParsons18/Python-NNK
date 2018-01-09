from openpyxl import Workbook
from openpyxl.worksheet import dimensions
import re

def writeVals(outerDict, innerDict, filename, board):
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', \
               'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    funcRegex = re.compile(r"(\^1.0)")
    
        
    wb = Workbook()
    ws = wb.active


    
    ws.title = 'Number Combinations'

    ws.merge_cells('B1:D1')
    ws['B1'] = 'National Number Knockout Number Combinations'

    ws['A4'] = 'Numbers'
    ws['B4'] = 'Unique Answers Found'
    ws['C4'] = 'Begin Formulas'

    outerDictKeys = sorted(outerDict.keys())
    row = 5
    for key in outerDictKeys:
        ws['A'+str(row)] = str(key)
        ws['B'+str(row)] = len(outerDict[key])
        col = 2
        mod = ''
        modNum = -1
        for ans in outerDict[key]:
            if col > 25:
                col = col - 26
                modNum += 1
                mod = columns[modNum]
                
                
        
            try:
                innerKey = key, ans
                for func in innerDict[innerKey]:
                    print func
                    func = funcRegex.sub('',func,0)
                    if mod != '':
                        ws[mod+columns[col]+str(row)] = '='+func
                        col += 1
                    else:
                        ws[columns[col]+str(row)] = '='+func
                        col += 1
                    
            except:
                continue
        row += 1

    ws2 = wb.create_sheet("Board")
    ws2.title = "Board"
    ws2["A1"] = "Board"
    row = 2
    col = 0
    board = sorted(board)
    for val in board:
        ws2[columns[col]+str(row)] = val
        col += 1
        if col > 10:
            row += 1
            col = 0
    try:
        wb.save(filename+'.xlsx')
        return True
    except:
        return 'Fail'
    
    
        
        
    

